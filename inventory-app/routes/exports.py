from flask import Blueprint, send_file
from flask_login import login_required
from io import BytesIO
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

exports_bp = Blueprint('exports', __name__)

HDR_FILL = PatternFill('solid', fgColor='2C3E50')
HDR_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_FILL = PatternFill('solid', fgColor='EBF5FB')
THIN = Side(style='thin', color='BDC3C7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _style_header(ws, headers, widths=None):
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        if widths and i <= len(widths):
            ws.column_dimensions[cell.column_letter].width = widths[i-1]


def _style_rows(ws, num_cols):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = LEFT
            if cell.row % 2 == 0:
                cell.fill = ALT_FILL


@exports_bp.route('/purchases')
@exports_bp.route('/purchase')
@login_required
def purchase_excel():
    from models import PurchaseMaster
    masters = PurchaseMaster.query.order_by(PurchaseMaster.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Purchase Records'
    ws.freeze_panes = 'A2'
    headers = ['#', 'Transaction ID', 'Vendor', 'Purchase Date', 'RV No', 'RV Date',
               'Item Name', 'Category', 'Qty', 'Unit Price', 'Line Total', 'Grand Total']
    _style_header(ws, headers, [5, 14, 22, 14, 14, 14, 26, 18, 8, 12, 14, 14])
    row = 2
    for m in masters:
        for j, pi in enumerate(m.purchase_items):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=m.transaction_id)
            ws.cell(row=row, column=3, value=pi.vendor.vendor_name if pi.vendor else (m.vendor.vendor_name if m.vendor else ''))
            ws.cell(row=row, column=4, value=m.purchase_date.strftime('%d %b %Y') if m.purchase_date else '')
            ws.cell(row=row, column=5, value=m.rv_no or '')
            ws.cell(row=row, column=6, value=m.rv_date.strftime('%d %b %Y') if m.rv_date else '')
            ws.cell(row=row, column=7, value=pi.item.item_name if pi.item else '')
            ws.cell(row=row, column=8, value=pi.item.category.category_name if pi.item and pi.item.category else '')
            ws.cell(row=row, column=9, value=pi.quantity)
            ws.cell(row=row, column=10, value=float(pi.unit_price or 0))
            ws.cell(row=row, column=11, value=float(pi.total_amount or 0))
            ws.cell(row=row, column=12, value=float(m.grand_total or 0) if j == 0 else '')
            row += 1
    _style_rows(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Purchase_Records_{date.today()}.xlsx')


@exports_bp.route('/distributions')
@exports_bp.route('/distribution')
@login_required
def distribution_excel():
    from models import DistributionMaster
    masters = DistributionMaster.query.order_by(DistributionMaster.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Distribution Records'
    ws.freeze_panes = 'A2'
    headers = ['#', 'Transaction ID', 'Voucher No', 'Issued To', 'Employee ID',
               'Phone', 'Department', 'Voucher Date', 'Item Name', 'Category', 'Qty Issued']
    _style_header(ws, headers, [5, 14, 16, 24, 14, 14, 18, 14, 26, 18, 10])
    row = 2
    for m in masters:
        for j, di in enumerate(m.dist_items):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=m.transaction_id)
            ws.cell(row=row, column=3, value=m.voucher_no or '')
            ws.cell(row=row, column=4, value=m.issued_to)
            ws.cell(row=row, column=5, value=m.employee_id or '')
            ws.cell(row=row, column=6, value=m.phone or '')
            ws.cell(row=row, column=7, value=m.department or '')
            ws.cell(row=row, column=8, value=m.voucher_date.strftime('%d %b %Y') if m.voucher_date else '')
            ws.cell(row=row, column=9, value=di.item.item_name if di.item else '')
            ws.cell(row=row, column=10, value=di.item.category.category_name if di.item and di.item.category else '')
            ws.cell(row=row, column=11, value=di.quantity_issued)
            row += 1
    _style_rows(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Distribution_Records_{date.today()}.xlsx')


@exports_bp.route('/warranty')
@login_required
def warranty_excel():
    from models import Item
    items = Item.query.filter(Item.warranty_expiry_date.isnot(None)).order_by(Item.warranty_expiry_date).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Warranty Report'
    ws.freeze_panes = 'A2'
    headers = ['#', 'Item Name', 'Category', 'Sub-Category', 'Brand', 'Model', 'Serial No', 'Warranty Expiry', 'Status']
    _style_header(ws, headers, [5, 26, 16, 16, 14, 18, 18, 16, 12])
    for i, it in enumerate(items, 1):
        expired = it.is_warranty_expired
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=it.item_name)
        ws.cell(row=i+1, column=3, value=it.category.category_name if it.category else '')
        ws.cell(row=i+1, column=4, value=it.subcategory.subcategory_name if it.subcategory else '')
        ws.cell(row=i+1, column=5, value=it.brand.brand_name if it.brand else '')
        ws.cell(row=i+1, column=6, value=it.model_name or '')
        ws.cell(row=i+1, column=7, value=it.serial_no or '')
        ws.cell(row=i+1, column=8, value=it.warranty_expiry_date.strftime('%d %b %Y') if it.warranty_expiry_date else '')
        ws.cell(row=i+1, column=9, value='Expired' if expired else 'Active')
        if expired:
            for c in range(1, 10):
                ws.cell(row=i+1, column=c).fill = PatternFill('solid', fgColor='FADBD8')
    _style_rows(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Warranty_Report_{date.today()}.xlsx')


@exports_bp.route('/stocks')
@login_required
def stocks_excel():
    from models import Item
    items = Item.query.order_by(Item.item_name).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Report'
    ws.freeze_panes = 'A2'
    headers = ['#', 'Item Name', 'Category', 'Sub-Category', 'Brand', 'Model', 'Serial No', 'Current Stock', 'Min Qty', 'Status']
    _style_header(ws, headers, [5, 26, 16, 16, 14, 18, 18, 14, 10, 12])
    for i, it in enumerate(items, 1):
        status = 'Out of Stock' if it.current_stock == 0 else ('Low Stock' if it.is_low_stock else 'OK')
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=it.item_name)
        ws.cell(row=i+1, column=3, value=it.category.category_name if it.category else '')
        ws.cell(row=i+1, column=4, value=it.subcategory.subcategory_name if it.subcategory else '')
        ws.cell(row=i+1, column=5, value=it.brand.brand_name if it.brand else '')
        ws.cell(row=i+1, column=6, value=it.model_name or '')
        ws.cell(row=i+1, column=7, value=it.serial_no or '')
        ws.cell(row=i+1, column=8, value=it.current_stock)
        ws.cell(row=i+1, column=9, value=it.minimum_stock_level)
        ws.cell(row=i+1, column=10, value=status)
        fill_color = 'FADBD8' if it.current_stock == 0 else ('FEF9E7' if it.is_low_stock else None)
        if fill_color:
            for c in range(1, 11):
                ws.cell(row=i+1, column=c).fill = PatternFill('solid', fgColor=fill_color)
    _style_rows(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Stock_Report_{date.today()}.xlsx')
